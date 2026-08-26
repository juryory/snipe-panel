"""Excel 批量导入 / 导出。

存量设备动辄几百台,一台台手输既慢又必然出错。这里的设计原则只有一条:
**先预演,再落库**。导入前把每一行的判定结果摊开给人看,确认无误才写。

导入是全有或全无的:只要有一行报错就整批拒绝。部分导入会让人搞不清到底进了
哪些、还差哪些,补录时更容易出乱子。
"""
import io
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from .models import STATUS_LABELS, Asset, AssetStatus, Category, Company, User

# Excel 列名 → 内部字段。与导出的表头一致,这样「导出 → 改 → 导入」能闭环
COLUMNS: List[Tuple[str, str]] = [
    ("资产编号", "asset_tag"),
    ("设备名称", "name"),
    ("分类", "category"),
    ("品牌", "brand"),
    ("型号", "model"),
    ("序列号", "serial_no"),
    ("状态", "status"),
    ("存放位置", "location"),
    ("长期责任人", "owner"),
    ("采购公司", "company"),
    ("采购日期", "purchased_at"),
    ("备注", "note"),
]

REQUIRED = ("name", "category")
LABEL_TO_STATUS = {label: value for value, label in STATUS_LABELS.items()}

TEMPLATE_NOTES = [
    "带 * 的列必填。",
    "资产编号留空则按分类前缀自动生成;只有导入存量设备、需要沿用旧编号时才填。",
    "分类必须是系统里已有的分类名称,不存在会报错(分类需要编号前缀,不能凭空创建)。",
    "长期责任人填用户名或姓名都可以;采购公司填公司名称。",
    "状态留空默认「在库」,可填「在库」「维修」「报废」。",
    "采购日期格式 2026-03-15,或直接用 Excel 的日期格式。",
]


class RowError(Exception):
    pass


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_date(value) -> Optional[date]:
    text = _text(value)
    if not text:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise RowError(f"采购日期「{text}」无法识别,请用 2026-03-15 这种格式")


def build_template() -> bytes:
    """生成导入模板。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "设备"

    headers = [f"{label} *" if field in REQUIRED else label for label, field in COLUMNS]
    ws.append(headers)
    ws.append(
        [
            "",
            "ThinkPad X1 Carbon",
            "电脑",
            "Lenovo",
            "Gen11",
            "PF3ABCDE",
            "在库",
            "库房 A",
            "张三",
            "星光影视器材",
            "2026-03-15",
            "行政批次",
        ]
    )
    for i, _ in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18

    notes = wb.create_sheet("填写说明")
    notes.column_dimensions["A"].width = 100
    for line in TEMPLATE_NOTES:
        notes.append([line])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _lookup_maps(db: Session) -> Dict[str, Dict[str, object]]:
    categories = db.execute(select(Category)).scalars().all()
    companies = db.execute(select(Company)).scalars().all()
    users = db.execute(select(User).where(User.is_active.is_(True))).scalars().all()

    users_by_key: Dict[str, List[User]] = {}
    for u in users:
        for key in filter(None, (u.username, u.real_name)):
            users_by_key.setdefault(key.strip(), []).append(u)

    return {
        "categories": {c.name.strip(): c for c in categories},
        "companies": {c.name.strip(): c for c in companies},
        "users": users_by_key,
    }


def parse(content: bytes) -> Tuple[List[Dict], List[str]]:
    """读表。返回(行数据, 整表级错误)。"""
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl 的异常类型很杂,统一兜住
        return [], [f"无法读取文件,请确认是 .xlsx 格式:{exc}"]

    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ["表格是空的"]

    # 表头允许带 * 和空格
    header = [_text(c).rstrip("*").strip() for c in rows[0]]
    label_to_field = {label: field for label, field in COLUMNS}
    index_of: Dict[str, int] = {}
    for i, label in enumerate(header):
        field = label_to_field.get(label)
        if field:
            index_of[field] = i

    missing = [label for label, field in COLUMNS if field in REQUIRED and field not in index_of]
    if missing:
        return [], [f"表头缺少必填列:{'、'.join(missing)}。建议先下载模板。"]

    parsed = []
    for line_no, raw in enumerate(rows[1:], start=2):
        values = {field: (raw[i] if i < len(raw) else None) for field, i in index_of.items()}
        if not any(_text(v) for v in values.values()):
            continue  # 整行空白,跳过
        parsed.append({"row": line_no, "values": values})
    return parsed, []


def validate(db: Session, parsed: List[Dict], create_missing_companies: bool) -> List[Dict]:
    """逐行判定。不写库,结果供预览和实际导入共用。"""
    maps = _lookup_maps(db)
    existing_tags = set(
        db.execute(select(Asset.asset_tag).where(Asset.deleted_at.is_(None))).scalars().all()
    )
    seen_tags: Dict[str, int] = {}
    seen_serials: Dict[str, int] = {}

    results = []
    for item in parsed:
        row, values = item["row"], item["values"]
        errors: List[str] = []
        warnings: List[str] = []
        data: Dict[str, object] = {}

        name = _text(values.get("name"))
        if not name:
            errors.append("设备名称不能为空")
        data["name"] = name

        cat_name = _text(values.get("category"))
        category = maps["categories"].get(cat_name)
        if not cat_name:
            errors.append("分类不能为空")
        elif category is None:
            errors.append(
                f"分类「{cat_name}」不存在。请先在系统里建好分类(分类需要编号前缀,不能自动创建)"
            )
        data["category"] = category

        tag = _text(values.get("asset_tag")).upper()
        if tag:
            if tag in existing_tags:
                errors.append(f"资产编号 {tag} 已存在")
            elif tag in seen_tags:
                errors.append(f"资产编号 {tag} 与第 {seen_tags[tag]} 行重复")
            else:
                seen_tags[tag] = row
        data["asset_tag"] = tag or None

        serial = _text(values.get("serial_no"))
        if serial:
            if serial in seen_serials:
                warnings.append(f"序列号与第 {seen_serials[serial]} 行相同")
            else:
                seen_serials[serial] = row
        data["serial_no"] = serial

        status_label = _text(values.get("status"))
        if status_label:
            status = LABEL_TO_STATUS.get(status_label)
            if status is None:
                errors.append(f"状态「{status_label}」无法识别,只能填:在库 / 维修 / 报废")
            data["status"] = status or AssetStatus.IN_STOCK
        else:
            data["status"] = AssetStatus.IN_STOCK

        owner_name = _text(values.get("owner"))
        owner = None
        if owner_name:
            hits = maps["users"].get(owner_name, [])
            if not hits:
                errors.append(f"找不到用户「{owner_name}」,请先创建该账号或留空")
            elif len(hits) > 1:
                errors.append(f"「{owner_name}」对应多个账号,请改填用户名")
            else:
                owner = hits[0]
        data["owner"] = owner

        company_name = _text(values.get("company"))
        company = None
        if company_name:
            company = maps["companies"].get(company_name)
            if company is None:
                if create_missing_companies:
                    warnings.append(f"将新建采购公司「{company_name}」")
                    data["new_company"] = company_name
                else:
                    errors.append(
                        f"采购公司「{company_name}」不存在。勾选「自动创建采购公司」或先手工建好"
                    )
        data["company"] = company

        try:
            data["purchased_at"] = _parse_date(values.get("purchased_at"))
        except RowError as exc:
            errors.append(str(exc))
            data["purchased_at"] = None

        data["brand"] = _text(values.get("brand"))
        data["model"] = _text(values.get("model"))
        data["location"] = _text(values.get("location"))
        data["note"] = _text(values.get("note"))

        results.append({"row": row, "data": data, "errors": errors, "warnings": warnings})
    return results


def export_workbook(assets: List[Asset], open_map: Dict[int, object]) -> bytes:
    """把台账导出成 xlsx。表头与导入模板一致,可以改完再导回来。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "设备台账"
    ws.append([label for label, _ in COLUMNS] + ["当前借用人"])

    for a in assets:
        checkout = open_map.get(a.id)
        borrower = ""
        if checkout is not None and checkout.user is not None:
            borrower = checkout.user.real_name or checkout.user.username
        ws.append(
            [
                a.asset_tag,
                a.name,
                a.category.name if a.category else "",
                a.brand,
                a.model,
                a.serial_no,
                STATUS_LABELS[a.status],
                a.location,
                (a.owner.real_name or a.owner.username) if a.owner else "",
                a.company.name if a.company else "",
                a.purchased_at.isoformat() if a.purchased_at else "",
                a.note,
                borrower,
            ]
        )

    for i in range(1, len(COLUMNS) + 2):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
