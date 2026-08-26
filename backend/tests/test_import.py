"""Excel 批量导入 / 导出。

重点在「预演不写库」和「有错就整批拒绝」——这两条是几百台设备敢往里灌的前提。
"""
import io

from openpyxl import Workbook, load_workbook

from tests.test_api import _activate, _make_asset, _new_user

HEADERS = [
    "资产编号", "设备名称", "分类", "品牌", "型号", "序列号",
    "状态", "存放位置", "长期责任人", "采购公司", "采购日期", "备注",
]


def _sheet(rows, headers=None):
    """把若干行拼成 xlsx 字节流。rows 里每项是与 HEADERS 等长的列表。"""
    wb = Workbook()
    ws = wb.active
    ws.append(headers if headers is not None else HEADERS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row(name, category="电脑", tag="", brand="", model="", serial="", status="",
         location="", owner="", company="", purchased="", note=""):
    return [tag, name, category, brand, model, serial, status, location, owner,
            company, purchased, note]


def _upload(client, url, content, **form):
    return client.post(
        url,
        files={"file": ("设备.xlsx", content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={k: str(v) for k, v in form.items()},
    )


# ---------- 模板 ----------
def test_template_downloads_and_matches_import_headers(admin):
    r = admin.get("/api/assets/import/template")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    header = [str(c.value).rstrip("*").strip() for c in wb[wb.sheetnames[0]][1]]
    assert header == HEADERS
    assert "填写说明" in wb.sheetnames


# ---------- 预演 ----------
def test_preview_does_not_write_anything(admin):
    content = _sheet([_row("电脑一"), _row("电脑二")])
    r = _upload(admin, "/api/assets/import/preview", content)
    assert r.status_code == 200, r.text
    data = r.json()
    assert (data["total"], data["ok_count"], data["error_count"]) == (2, 2, 0)
    assert data["committed"] is False
    # 关键:预演之后台账里必须一台都没有
    assert admin.get("/api/assets").json()["total"] == 0


def test_preview_reports_row_numbers_matching_excel(admin):
    """报错要带 Excel 行号,否则几百行里根本找不到是哪一行。"""
    content = _sheet([_row("好的"), _row("坏的", category="不存在的分类")])
    rows = _upload(admin, "/api/assets/import/preview", content).json()["rows"]
    assert [r["row"] for r in rows] == [2, 3]  # 第 1 行是表头
    assert rows[1]["ok"] is False
    assert "不存在" in rows[1]["errors"][0]


def test_unknown_category_is_rejected_not_auto_created(admin):
    """分类需要编号前缀,不能凭空创建 —— 打错字就该报错,而不是造出个垃圾分类。"""
    content = _sheet([_row("设备", category="相機")])  # 繁体,不匹配
    data = _upload(admin, "/api/assets/import/preview", content).json()
    assert data["error_count"] == 1
    assert admin.get("/api/categories").json().__len__() == 5  # 还是默认那 5 个


# ---------- 实际导入 ----------
def test_import_creates_assets_and_assigns_tags(admin):
    content = _sheet([
        _row("电脑一", brand="Lenovo", location="库房 A"),
        _row("电脑二"),
        _row("相机一", category="相机"),
    ])
    r = _upload(admin, "/api/assets/import", content)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["committed"] is True
    assert data["ok_count"] == 3

    listed = admin.get("/api/assets").json()
    assert listed["total"] == 3
    tags = sorted(a["asset_tag"] for a in listed["items"])
    assert tags == ["CAM-0001", "PC-0001", "PC-0002"]
    # 返回结果里要带上实际分到的编号,方便随后去打标签
    assert all(row["asset_tag"] for row in data["rows"])


def test_import_is_all_or_nothing(admin):
    """一行坏的就整批拒绝。部分导入会让人搞不清进了哪些、还差哪些。"""
    content = _sheet([_row("好的一"), _row("坏的", category="没有这个"), _row("好的二")])
    r = _upload(admin, "/api/assets/import", content)
    assert r.status_code == 422
    assert "已全部取消" in r.json()["detail"]
    assert admin.get("/api/assets").json()["total"] == 0


def test_explicit_tags_are_honoured_for_legacy_assets(admin):
    content = _sheet([_row("存量一", tag="PC-8001"), _row("存量二", tag="PC-8002")])
    assert _upload(admin, "/api/assets/import", content).status_code == 200
    tags = {a["asset_tag"] for a in admin.get("/api/assets").json()["items"]}
    assert tags == {"PC-8001", "PC-8002"}


def test_duplicate_tag_inside_the_file_is_caught(admin):
    content = _sheet([_row("一", tag="PC-9001"), _row("二", tag="PC-9001")])
    rows = _upload(admin, "/api/assets/import/preview", content).json()["rows"]
    assert rows[1]["ok"] is False
    assert "第 2 行重复" in rows[1]["errors"][0]


def test_tag_colliding_with_existing_asset_is_caught(admin):
    _make_asset(admin, "已有的", tag="PC-7001")
    content = _sheet([_row("新的", tag="PC-7001")])
    rows = _upload(admin, "/api/assets/import/preview", content).json()["rows"]
    assert rows[0]["ok"] is False
    assert "已存在" in rows[0]["errors"][0]


def test_duplicate_serial_is_only_a_warning(admin):
    """SN 重复多半是抄串了,但也可能是厂商真给了一样的号 —— 提醒但不拦。"""
    content = _sheet([_row("一", serial="SN123"), _row("二", serial="SN123")])
    data = _upload(admin, "/api/assets/import/preview", content).json()
    assert data["error_count"] == 0
    assert data["rows"][1]["warnings"]


# ---------- 关联字段 ----------
def test_owner_matches_by_username_or_real_name(admin):
    _new_user(admin, "alice")  # real_name 是 ALICE
    content = _sheet([_row("按用户名", owner="alice"), _row("按姓名", owner="ALICE")])
    assert _upload(admin, "/api/assets/import", content).status_code == 200
    owners = {a["name"]: a["owner"]["username"] for a in admin.get("/api/assets").json()["items"]}
    assert owners == {"按用户名": "alice", "按姓名": "alice"}


def test_unknown_owner_is_rejected(admin):
    content = _sheet([_row("设备", owner="查无此人")])
    rows = _upload(admin, "/api/assets/import/preview", content).json()["rows"]
    assert rows[0]["ok"] is False
    assert "找不到用户" in rows[0]["errors"][0]


def test_company_must_exist_unless_auto_create_is_on(admin):
    content = _sheet([_row("设备", company="新公司")])

    rows = _upload(admin, "/api/assets/import/preview", content).json()["rows"]
    assert rows[0]["ok"] is False

    r = _upload(admin, "/api/assets/import", content, create_missing_companies=True)
    assert r.status_code == 200, r.text
    assert [c["name"] for c in admin.get("/api/companies").json()] == ["新公司"]


def test_same_new_company_on_many_rows_is_created_once(admin):
    content = _sheet([_row("一", company="同一家"), _row("二", company="同一家")])
    assert _upload(admin, "/api/assets/import", content, create_missing_companies=True).status_code == 200
    companies = admin.get("/api/companies").json()
    assert len(companies) == 1
    assert companies[0]["asset_count"] == 2


def test_status_and_date_are_parsed(admin):
    content = _sheet([
        _row("维修中的", status="维修", purchased="2026-03-15"),
        _row("斜杠日期", purchased="2026/04/01"),
    ])
    assert _upload(admin, "/api/assets/import", content).status_code == 200
    by_name = {a["name"]: a for a in admin.get("/api/assets").json()["items"]}
    assert by_name["维修中的"]["status"] == "repair"
    assert by_name["维修中的"]["purchased_at"] == "2026-03-15"
    assert by_name["斜杠日期"]["purchased_at"] == "2026-04-01"


def test_bad_status_and_date_are_rejected(admin):
    content = _sheet([_row("一", status="借出"), _row("二", purchased="去年")])
    rows = _upload(admin, "/api/assets/import/preview", content).json()["rows"]
    assert "无法识别" in rows[0]["errors"][0]
    assert "无法识别" in rows[1]["errors"][0]


# ---------- 表格本身的问题 ----------
def test_missing_required_column_is_reported_upfront(admin):
    bad = _sheet([["PC-0001", "只有编号和名称"]], headers=["资产编号", "设备名称"])
    r = _upload(admin, "/api/assets/import/preview", bad)
    assert r.status_code == 400
    assert "分类" in r.json()["detail"]


def test_blank_rows_are_skipped(admin):
    content = _sheet([_row("一"), ["", "", "", "", "", "", "", "", "", "", "", ""], _row("二")])
    data = _upload(admin, "/api/assets/import/preview", content).json()
    assert data["total"] == 2


def test_non_excel_upload_is_rejected(admin):
    r = _upload(admin, "/api/assets/import/preview", b"this is not a spreadsheet")
    assert r.status_code == 400
    assert "xlsx" in r.json()["detail"]


def test_regular_user_cannot_import(admin):
    _new_user(admin, "alice")
    alice = _activate("alice")
    r = _upload(alice, "/api/assets/import/preview", _sheet([_row("设备")]))
    assert r.status_code == 403


# ---------- 导出 ----------
def test_export_round_trips_back_into_import(admin):
    """导出的表头必须能直接被导入吃掉,否则「导出改完再导回」就断了。"""
    _make_asset(admin, "原有设备")
    r = admin.get("/api/assets/export")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb[wb.sheetnames[0]]
    header = [str(c.value) for c in ws[1]]
    assert header[: len(HEADERS)] == HEADERS

    # 把导出的内容原样喂回预览:编号已存在,所以每行都该报「已存在」而不是格式错误
    rows = _upload(admin, "/api/assets/import/preview", r.content).json()["rows"]
    assert len(rows) == 1
    assert "已存在" in rows[0]["errors"][0]


def test_export_respects_filters(admin):
    _make_asset(admin, "电脑")
    admin.post("/api/assets", json={
        "name": "相机",
        "category_id": next(c["id"] for c in admin.get("/api/categories").json()
                            if c["tag_prefix"] == "CAM"),
    })
    r = admin.get("/api/assets/export", params={"q": "相机"})
    ws = load_workbook(io.BytesIO(r.content))[wb_name(r)]
    assert ws.max_row == 2  # 表头 + 1 行


def wb_name(response):
    return load_workbook(io.BytesIO(response.content)).sheetnames[0]
